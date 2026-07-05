#!/usr/bin/env python3
"""
Phase 3: 题库 Sheet 生成 — 解析整合内容(v2) 的 ## / ### / - 三级结构，拍平为行式数据

用法:
  python3 generate_bank.py [--xlsx 笔记列表_含OCR.xlsx]

配置:
  export BANK_XLSX="/path/to/excel.xlsx"
"""

import os, re, sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment

XLSX = os.getenv("BANK_XLSX", "笔记列表_含OCR.xlsx")
SHEET_NAME = "面试题库"
V2_COL = 28    # 整合内容(v2)
TITLE_COL = 5
LINK_COL = 17
DATE_COL = 15
ID_COL = 1

COMPANIES = [
    "滴滴", "字节", "腾讯", "美团", "快手", "百度", "蚂蚁", "阿里",
    "京东", "拼多多", "携程", "得物", "小红书", "B站", "网易",
]
POSITION_KEYWORDS = [
    ("Java", ["java"]),
    ("前端", ["前端"]),
    ("算法", ["算法", "推荐", "搜索"]),
    ("数据", ["数据开发", "大数据", "数仓", "数据分析"]),
    ("客户端", ["客户端"]),
    ("AI/大模型", ["大模型", "agent", "rag"]),
    ("Go", ["go", "golang"]),
]

SKIP_LINE = re.compile(
    r"^(?:"
    r"[（(]\d+[min分钟）)]+[\u3000-\u303F\uff00-\uffef]*"
    r"|(?:准备建议|总结|整体难度|耗时|一面|二面|三面|四面|五面)[：:\s\d]*min.*"
    r"|注意[：:]\s*"
    r"|[：:]\s*$"
    r"|^(简洁干练|条理清晰|避免夸大|全程围绕|项目极致深挖|洽就是)"
    r")", re.IGNORECASE)


def is_sql_continuation(line):
    stripped = line.strip()
    if not stripped:
        return False
    if stripped.startswith("."):
        return True
    if re.match(r"^(create|insert|select|drop|alter|update|delete|from|where|set|into|values|join|on|group|order|having|limit|union)\b", stripped, re.IGNORECASE):
        return True
    if stripped.endswith(";"):
        return True
    if re.match(r"^[a-z.&|(]", stripped):
        return True
    return False


def extract_company(title):
    for c in COMPANIES:
        if c in title:
            return c
    return ""


def extract_position(title, content):
    text = (title + " " + content).lower()
    for pos, keywords in POSITION_KEYWORDS:
        for kw in keywords:
            if kw in text:
                return pos
    return ""


def main():
    wb = load_workbook(XLSX)
    ws = wb["笔记列表"]

    # 如果已有面试题库 Sheet，替换
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws2 = wb.create_sheet(SHEET_NAME)

    headers = ["笔记ID", "标题", "公司", "岗位", "轮次", "技术领域", "题目", "链接", "发布时间"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)

    questions = []
    stats = {
        "extracted": 0, "skipped_empty": 0, "skipped_discussion": 0,
        "skipped_noise": 0, "merged_sql": 0, "skipped_short": 0,
        "skipped_area": 0,
    }

    for r in range(2, ws.max_row + 1):
        title = str(ws.cell(row=r, column=TITLE_COL).value or "")
        v2 = str(ws.cell(row=r, column=V2_COL).value or "")
        link = str(ws.cell(row=r, column=LINK_COL).value or "")
        pub_date = str(ws.cell(row=r, column=DATE_COL).value or "")
        note_id = str(ws.cell(row=r, column=ID_COL).value or "")

        if not v2.strip():
            stats["skipped_empty"] += 1
            continue
        if v2 == "（该笔记为行业讨论帖）" or v2 == "（该笔记为空）":
            stats["skipped_discussion"] += 1
            continue

        company = extract_company(title)
        position = extract_position(title, v2)
        current_round = ""
        current_area = "其他"

        lines = v2.split("\n")
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line.startswith("## "):
                current_round = re.sub(r"^##\s+", "", line)
                i += 1
                continue
            if line.startswith("### "):
                current_area = re.sub(r"^###\s+", "", line)
                i += 1
                continue
            if line.startswith("- "):
                q = re.sub(r"^-\s+", "", line)

                if SKIP_LINE.match(q):
                    stats["skipped_noise"] += 1
                    i += 1
                    continue
                if len(q) < 5:
                    stats["skipped_short"] += 1
                    i += 1
                    continue

                # SQL 续行合并
                merged_q = q
                j = i + 1
                while j < len(lines):
                    next_line = lines[j].strip()
                    if next_line.startswith("- "):
                        next_q = re.sub(r"^-\s+", "", next_line)
                        if is_sql_continuation(next_q):
                            merged_q += "\n" + next_q
                            stats["merged_sql"] += 1
                            j += 1
                            continue
                        break
                    elif next_line == "":
                        j += 1
                        continue
                    else:
                        break

                questions.append((note_id, title, company, position,
                                  current_round, current_area, merged_q, link, pub_date))
                stats["extracted"] += 1
                i = j
                continue
            i += 1

        if r % 20 == 0:
            print(f"  已扫描 {r-1}/{ws.max_row-1} 条笔记，提取 {stats['extracted']} 道题...")

    row_idx = 2
    for q in questions:
        for col, val in enumerate(q, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            if col in (7, 8):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1

    widths = {"A": 12, "B": 35, "C": 10, "D": 10, "E": 15, "F": 12, "G": 60, "H": 40, "I": 20}
    for col_letter, w in widths.items():
        ws2.column_dimensions[col_letter].width = w

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:I{row_idx - 1}"

    wb.save(XLSX)

    print(f"\n生成完成！")
    print(f"  提取题目: {stats['extracted']} 道")
    print(f"  合并 SQL 续行: {stats['merged_sql']}")
    print(f"  跳过空笔记: {stats['skipped_empty']}")
    print(f"  跳过讨论帖: {stats['skipped_discussion']}")
    print(f"  跳过噪声: {stats['skipped_noise']}")
    print(f"  跳过过短: {stats['skipped_short']}")
    print(f"  Sheet: {SHEET_NAME}")


if __name__ == "__main__":
    main()
