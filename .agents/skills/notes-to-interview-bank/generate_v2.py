#!/usr/bin/env python3
"""
Phase 2: 逐条 LLM 结构化 — 将 正文+OCR整合 转换为 ## 面试题 / ### 领域 / - 问题 格式

配置方式（优先级：环境变量 > 脚本默认值）:
  export V2_API_URL="https://api.openai.com/v1/chat/completions"
  export V2_API_KEY="sk-..."
  export V2_MODEL="gpt-4o-mini"
  export V2_PROTOCOL="openai"        # "openai" 或 "anthropic"
  export V2_XLSX="/path/to/excel.xlsx"
"""

import json, os, time, sys
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from openpyxl import load_workbook

# ============ 配置 ============
DEFAULT_API_URL = os.getenv("V2_API_URL", "https://api.openai.com/v1/chat/completions")
DEFAULT_API_KEY = os.getenv("V2_API_KEY", "")
DEFAULT_MODEL = os.getenv("V2_MODEL", "gpt-4o-mini")
DEFAULT_PROTOCOL = os.getenv("V2_PROTOCOL", "openai")  # "openai" | "anthropic"
XLSX = os.getenv("V2_XLSX", "笔记列表_含OCR.xlsx")

SOURCE_COL = 26   # 正文+OCR整合
TARGET_COL = 28   # 整合内容(v2)

SKIP_IDS = set()  # 手动修复的笔记ID，需要时填充

PROMPT_TEMPLATE = """## 角色
你是一名面试题库整理专家。

## 任务
下面是一条面经笔记的原始内容。请提取所有面试题，按技术领域分组，输出为以下格式：

## 面试题
### Java
- [完整的问题描述]
- [完整的问题描述]
### MySQL
- [完整的问题描述]
...

## 要求
1. 问题表述务必完整，不能截断
2. 保留 SQL 代码、算法题描述等细节
3. 根据内容判断技术领域（Java/MySQL/Redis/MQ/算法/系统设计/项目/网络/操作系统等）
4. 如果内容不含任何面试题（如招聘广告、闲聊、行业讨论），只输出"（该笔记为行业讨论帖）"
5. 只输出整理后的内容，不要额外解释

## 原始内容
{content}"""


def call_api_openai(prompt, retries=3):
    body = json.dumps({
        "model": DEFAULT_MODEL, "max_tokens": 4096, "stream": True,
        "messages": [{"role": "user", "content": prompt}]
    }).encode()
    req = Request(DEFAULT_API_URL, data=body, headers={
        "Content-Type": "application/json",
        "Authorization": f"Bearer {DEFAULT_API_KEY}",
    })
    for attempt in range(retries + 1):
        try:
            resp = urlopen(req, timeout=120)
            full = ""
            buf = b""
            while True:
                chunk = resp.read(4096)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    line = line.decode().strip()
                    if not line or not line.startswith("data: "):
                        continue
                    d = line[6:]
                    if d == "[DONE]":
                        continue
                    try:
                        data = json.loads(d)
                        for c in data.get("choices", []):
                            if c["delta"].get("content"):
                                full += c["delta"]["content"]
                    except json.JSONDecodeError:
                        pass
            return full.strip() or None
        except Exception as e:
            if attempt < retries:
                wait = 5 * (attempt + 1)
                print(f"  重试 {attempt+1}/{retries} (等待{wait}s)... ({type(e).__name__})")
                time.sleep(wait)
            else:
                print(f"  失败: {e}")
                return None


def call_api_anthropic(prompt, retries=3):
    api_url = DEFAULT_API_URL.rstrip("/")
    if not api_url.endswith("/messages"):
        api_url = api_url.replace("/chat/completions", "/messages")

    body = json.dumps({
        "model": DEFAULT_MODEL, "max_tokens": 4096,
        "messages": [{"role": "user", "content": prompt}]
    }).encode()
    req = Request(api_url, data=body, headers={
        "Content-Type": "application/json",
        "x-api-key": DEFAULT_API_KEY,
        "anthropic-version": "2023-06-01",
    })
    for attempt in range(retries + 1):
        try:
            resp = urlopen(req, timeout=120)
            full = ""
            buf = b""
            while True:
                chunk = resp.read(4096)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    line = line.decode().strip()
                    if not line or line.startswith(":"):
                        continue
                    if line.startswith("data: "):
                        d = line[6:]
                        if d == "[DONE]":
                            continue
                        try:
                            data = json.loads(d)
                            if data.get("type") == "content_block_delta":
                                full += data["delta"].get("text", "")
                            elif data.get("type") == "message_start":
                                for b in data["message"].get("content", []):
                                    if b.get("type") == "text":
                                        full += b.get("text", "")
                        except json.JSONDecodeError:
                            pass
            return full.strip() or None
        except Exception as e:
            if attempt < retries:
                wait = 5 * (attempt + 1)
                print(f"  重试 {attempt+1}/{retries} (等待{wait}s)... ({type(e).__name__})")
                time.sleep(wait)
            else:
                print(f"  失败: {e}")
                return None


def main():
    if not DEFAULT_API_KEY:
        print("错误：未设置 V2_API_KEY")
        sys.exit(1)

    call_api = call_api_openai if DEFAULT_PROTOCOL == "openai" else call_api_anthropic

    wb = load_workbook(XLSX)
    ws = wb["笔记列表"]
    total = ws.max_row - 1
    stats = {"success": 0, "skipped": 0, "empty": 0, "discussion": 0, "failed": 0}

    print(f"共 {total} 条笔记，协议={DEFAULT_PROTOCOL}，模型={DEFAULT_MODEL}\n")

    for r in range(2, ws.max_row + 1):
        nid = str(ws.cell(row=r, column=1).value or "")
        title = str(ws.cell(row=r, column=5).value or "")
        source = str(ws.cell(row=r, column=SOURCE_COL).value or "").strip()
        idx = r - 1

        print(f"[{idx}/{total}] {str(title)[:36]:36s}  ", end="", flush=True)

        if nid in SKIP_IDS:
            stats["skipped"] += 1
            print("SKIP")
            continue

        if not source:
            ws.cell(row=r, column=TARGET_COL).value = "（该笔记为空）"
            stats["empty"] += 1
            print("EMPTY")
            continue

        prompt = PROMPT_TEMPLATE.format(content=source)
        result = call_api(prompt)

        if result is None:
            stats["failed"] += 1
            print("FAIL")
            continue

        ws.cell(row=r, column=TARGET_COL).value = result

        if result == "（该笔记为行业讨论帖）":
            stats["discussion"] += 1
            print("DISCUSSION")
        else:
            stats["success"] += 1
            print("OK")

        if (r - 2) % 10 == 9:
            wb.save(XLSX)
            print(f"  → 已保存 ({idx}/{total})")

    wb.save(XLSX)
    print(f"\n完成! 成功={stats['success']} 跳过={stats['skipped']} 空={stats['empty']} 讨论={stats['discussion']} 失败={stats['failed']}")


if __name__ == "__main__":
    main()
